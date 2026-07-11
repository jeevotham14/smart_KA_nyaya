"""Save registered user details to an Excel sheet."""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Excel file lives next to the SQLite DB in the backend root
EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "registered_users.xlsx"

HEADERS = [
    "User ID",
    "Name",
    "Email",
    "Phone",
    "Language Preference",
    "District",
    "Taluk",
    "Role",
    "Registered At",
]


def _get_or_create_workbook() -> Workbook:
    """Load the workbook or create a fresh one with headers."""
    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registered Users"
    ws.append(HEADERS)

    # Style the header row
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Auto-width columns
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(len(header) + 4, 18)

    wb.save(EXCEL_PATH)
    return wb


def save_user_to_excel(user) -> None:
    """Append a single user row to the Excel file.

    `user` is expected to be a SQLAlchemy User model instance.
    """
    wb = _get_or_create_workbook()
    ws = wb.active

    ws.append([
        str(user.user_id),
        user.name,
        user.email,
        user.phone or "",
        user.language_pref,
        user.district or "",
        user.taluk or "",
        user.role,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])

    wb.save(EXCEL_PATH)
